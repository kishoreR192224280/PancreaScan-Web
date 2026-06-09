"""
PancreaScan E2E Test Report Generator
Runs all Selenium tests and produces a color-coded Excel report
matching the format of the reference sample report.
"""
import os
import time
import subprocess
import datetime
import xml.etree.ElementTree as ET

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Config ──────────────────────────────────────────────────────────────────────
TEST_DIR   = "tests/e2e"
XML_OUTPUT = "report.xml"
REPORT_DIR = "."


def run_pytest():
    """Run the full E2E test suite and return start/end timestamps."""
    print("\n[START] Starting PancreaScan E2E Test Suite...\n")
    start = datetime.datetime.now(datetime.timezone.utc)
    subprocess.run(
        ["python", "-m", "pytest", TEST_DIR, "-v", "--tb=short", f"--junitxml={XML_OUTPUT}"],
        cwd=os.getcwd()
    )
    end = datetime.datetime.now(datetime.timezone.utc)
    print("\n[DONE] Tests completed. Generating report...\n")
    return start, end


def parse_results(xml_path):
    """Parse the JUnit XML output from pytest into structured data."""
    tree = ET.parse(xml_path)
    root = tree.getroot()

    # Handle both <testsuites><testsuite> and <testsuite> structures
    suite = root.find("testsuite") or root

    total    = int(suite.get("tests",    0))
    failures = int(suite.get("failures", 0))
    errors   = int(suite.get("errors",   0))
    failed   = failures + errors
    passed   = total - failed
    duration = float(suite.get("time", 0.0))

    passed_rows  = []
    failed_rows  = []
    details_rows = []
    log_rows     = []

    p_no = f_no = d_no = 1

    # Human-readable category map from module prefix
    category_map = {
        "test_01": "Landing Page",
        "test_02": "Login Page",
        "test_03": "Register Page",
        "test_04": "Forgot Password",
        "test_05": "Dashboard Navigation",
        "test_06": "Dashboard Stats",
        "test_07": "CT Scan Upload & Workspace",
        "test_08": "Scan Report & PDF",
        "test_09": "Patient History",
        "test_10": "Analytics Tab",
        "test_11": "Settings Tab",
        "test_12": "Logout",
        "test_auth":     "Authentication (Legacy)",
        "test_dashboard": "Dashboard (Legacy)",
        "test_landing":  "Landing (Legacy)",
        "test_upload":   "Upload (Legacy)",
    }

    ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for case in suite.findall("testcase"):
        raw_class = case.get("classname", "")
        test_name = case.get("name", "")
        time_sec  = float(case.get("time", "0"))

        # Derive module from classname (e.g. "tests.e2e.test_02_login.TestLoginPageElements")
        parts = raw_class.split(".")
        module = next((p for p in parts if p.startswith("test_")), raw_class)
        prefix = "_".join(module.split("_")[:2])  # e.g. "test_02"
        category = category_map.get(prefix, module)
        full_name = f"{category} → {test_name}"

        failure = case.find("failure")
        error   = case.find("error")
        issue   = failure if failure is not None else error

        if issue is not None:
            raw_msg    = issue.get("message", "Test Failed")
            raw_detail = (issue.text or raw_msg).strip()

            # Make the error message human-readable
            human_msg = make_human_readable(raw_msg, test_name)
            issue_detail = raw_detail[:800]  # cap to avoid cell overflow

            failed_rows.append({
                "No."       : f_no,
                "Category"  : category,
                "Test Name" : test_name,
                "Error"     : human_msg,
                "Status"    : "FAILED",
                "Timestamp" : ts,
            })
            details_rows.append({
                "No."          : d_no,
                "Category"     : category,
                "Test Name"    : test_name,
                "Status"       : "FAILED",
                "Error Details": issue_detail,
            })
            log_rows.append({
                "Timestamp": ts,
                "Level"    : "ERROR",
                "Message"  : f"[{category}] {test_name} → FAILED: {human_msg}",
            })
            f_no += 1
        else:
            passed_rows.append({
                "No."       : p_no,
                "Category"  : category,
                "Test Name" : test_name,
                "Time (sec)": round(time_sec, 2),
                "Status"    : "PASSED",
            })
            details_rows.append({
                "No."          : d_no,
                "Category"     : category,
                "Test Name"    : test_name,
                "Status"       : "PASSED",
                "Error Details": "None — test passed successfully.",
            })
            log_rows.append({
                "Timestamp": ts,
                "Level"    : "INFO",
                "Message"  : f"[{category}] {test_name} → PASSED in {round(time_sec,2)}s",
            })
            p_no += 1
        d_no += 1

    return total, passed, failed, duration, passed_rows, failed_rows, details_rows, log_rows


def make_human_readable(raw_msg: str, test_name: str) -> str:
    """Convert cryptic Selenium/pytest error messages into plain English."""
    msg = raw_msg.lower()

    if "timeout" in msg or "timeoutexception" in msg:
        return (
            f"The test timed out waiting for an element to appear. "
            f"The page element expected by '{test_name}' did not become visible within the allowed time. "
            "This could mean the UI did not load, the element selector is wrong, or the app is slow."
        )
    if "nosuchelement" in msg:
        return (
            f"A required UI element was not found on the page during '{test_name}'. "
            "The element may have been removed, renamed, or is inside a shadow DOM. "
            "Check the CSS class or XPath selector used in this test."
        )
    if "stale element" in msg or "staleelementreference" in msg:
        return (
            f"A page element was found initially but became detached from the DOM before the test could interact with it. "
            f"This happened in '{test_name}'. It usually occurs when React re-renders the component mid-test. "
            "The fix is to re-locate the element after any page state change."
        )
    if "nosuchwindow" in msg or "web view not found" in msg:
        return (
            f"The browser window or tab closed unexpectedly during '{test_name}'. "
            "This typically happens when a previous test caused the browser to crash or the session was lost. "
            "Ensure the test has a valid open browser window before executing."
        )
    if "assert" in msg or "assertionerror" in msg:
        # Clean up the assertion message
        clean = raw_msg.replace("AssertionError:", "").strip()
        return clean if clean else f"An assertion check failed in '{test_name}'. The expected content was not found on the page."
    if "connectionrefused" in msg or "err_connection" in msg:
        return (
            f"The test could not connect to the website at all during '{test_name}'. "
            "The site may be offline, or the URL in the test configuration is wrong."
        )
    # Fallback — return a cleaned version
    return raw_msg[:200] if raw_msg else f"Unknown error in '{test_name}'."


def style_workbook(path):
    """Apply color coding, bold headers, and auto-width columns to the Excel file."""
    wb = load_workbook(path)

    GREEN   = PatternFill("solid", fgColor="C6EFCE")
    RED     = PatternFill("solid", fgColor="FFC7CE")
    HEADER  = PatternFill("solid", fgColor="1F3864")
    SUMMARY = PatternFill("solid", fgColor="2E4057")
    WHITE   = Font(color="FFFFFF", bold=True)
    BOLD    = Font(bold=True)
    THIN    = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Style header row
        for cell in ws[1]:
            cell.fill      = HEADER
            cell.font      = WHITE
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = THIN

        # Color-code data rows based on Status column
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border    = THIN
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            # Find status column
            for cell in row:
                if cell.value in ("PASSED", "INFO"):
                    for c in row: c.fill = GREEN
                    break
                elif cell.value in ("FAILED", "ERROR"):
                    for c in row: c.fill = RED
                    break

        # Auto-width columns (capped at 80 chars)
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    val_len = len(str(cell.value)) if cell.value else 0
                    max_len = max(max_len, val_len)
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 4, 80)

        # Freeze header row
        ws.freeze_panes = "A2"

    wb.save(path)


def generate_excel(start, end, total, passed, failed, duration,
                   passed_rows, failed_rows, details_rows, log_rows):
    """Build and save the final Excel report."""
    pass_rate = round((passed / total) * 100, 2) if total > 0 else 0

    summary = [{
        "Test Suite"    : "PancreaScan Web App — Full E2E Workflow",
        "Total Tests"   : total,
        "Passed"        : passed,
        "Failed"        : failed,
        "Pass Rate %"   : pass_rate,
        "Duration (sec)": round(duration, 2),
        "Start Time"    : start.strftime("%Y-%m-%dT%H:%M:%S.%fZ"),
        "End Time"      : end.strftime("%Y-%m-%dT%H:%M:%S.%fZ"),
    }]

    ts_str   = start.strftime("%Y-%m-%dT%H-%M-%S")
    filename = f"E2E_Test_Report_PancreaScan_{ts_str}.xlsx"
    filepath = os.path.join(REPORT_DIR, filename)

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        pd.DataFrame(summary).to_excel(writer,     sheet_name="Summary",       index=False)
        pd.DataFrame(passed_rows).to_excel(writer,  sheet_name="Passed Tests",  index=False)
        pd.DataFrame(failed_rows).to_excel(writer,  sheet_name="Failed Tests",  index=False)
        pd.DataFrame(log_rows).to_excel(writer,     sheet_name="Execution Log", index=False)
        pd.DataFrame(details_rows).to_excel(writer, sheet_name="Test Details",  index=False)

    style_workbook(filepath)
    return filepath, pass_rate


if __name__ == "__main__":
    start, end = run_pytest()

    if not os.path.exists(XML_OUTPUT):
        print("❌ report.xml not found. Tests may have crashed before generating output.")
        exit(1)

    total, passed, failed, duration, p_rows, f_rows, d_rows, l_rows = parse_results(XML_OUTPUT)
    filepath, pass_rate = generate_excel(start, end, total, passed, failed, duration, p_rows, f_rows, d_rows, l_rows)

    print(f"\n{'='*60}")
    print(f"  [REPORT] PancreaScan E2E Report Generated!")
    print(f"{'='*60}")
    print(f"  Total Tests : {total}")
    print(f"  [PASS]  Passed   : {passed}")
    print(f"  [FAIL]  Failed   : {failed}")
    print(f"  Pass Rate   : {pass_rate}%")
    print(f"  Duration    : {round(duration,1)}s")
    print(f"  Report      : {filepath}")
    print(f"{'='*60}\n")
