import json
import requests
import time
import os
from datetime import datetime, timezone
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INPUT_FILE = "input.json"
REPORT_FILE = "report.json"

ENDPOINTS = [
    # auth.php
    {"path": "/auth.php", "method": "POST", "data": {"action": "signup", "name": "Test User", "email": "testsignup@example.com", "password": "Password123!"}, "requires_auth": False},
    {"path": "/auth.php", "method": "POST", "data": {"action": "verify_otp", "email": "testsignup@example.com", "otp": "123456"}, "requires_auth": False},
    {"path": "/auth.php", "method": "POST", "data": {"action": "request_password_reset", "email": "testsignup@example.com"}, "requires_auth": False},
    {"path": "/auth.php", "method": "POST", "data": {"action": "reset_password", "email": "testsignup@example.com", "otp": "123456", "new_password": "NewPassword123!"}, "requires_auth": False},
    {"path": "/auth.php", "method": "POST", "data": {"action": "login", "email": "testsignup@example.com", "password": "Password123!"}, "requires_auth": False},
    {"path": "/auth.php", "method": "POST", "data": {"action": "delete_account"}, "requires_auth": True}, # Needs user_email conceptually, but script uses email
    
    # sync.php
    {"path": "/sync.php", "method": "POST", "data": {"action": "upload_scan", "image": "base64dummy"}, "requires_auth": True},
    {"path": "/sync.php", "method": "POST", "data": {"action": "get_history"}, "requires_auth": True},
    {"path": "/sync.php", "method": "POST", "data": {"action": "clear_history"}, "requires_auth": True},
    {"path": "/sync.php", "method": "POST", "data": {"action": "delete_scan", "timestamp": "2024-01-01 00:00:00"}, "requires_auth": True},
    
    # fl.php
    {"path": "/fl.php", "method": "POST", "data": {"action": "upload_gradients", "gradients": "{}"}, "requires_auth": True}, # Assume requires client_id/auth
    {"path": "/fl.php", "method": "POST", "data": {"action": "get_global_model"}, "requires_auth": False}, # Usually public for clients to download
]

def load_input():
    if not os.path.exists(INPUT_FILE):
        print(f"Error: {INPUT_FILE} not found.")
        return None
    with open(INPUT_FILE, "r") as f:
        return json.load(f)

def run_tests():
    config = load_input()
    if not config:
        return

    base_url = config.get("baseUrl")
    auth_email = config.get("doctor") # The backend uses email as an identifier in POST body, not headers
    
    report = []
    
    # Step 2: Expectation Model
    print("=== EXPECTATION MODEL ===")
    for ep in ENDPOINTS:
        access = "Requires Auth (via email param)" if ep['requires_auth'] else "Public"
        print(f"Endpoint: {ep['path']} | Action: {ep['data'].get('action')} | Access: {access}")
    print("=========================\n")

    # Step 3: Run Tests
    print("=== RUNNING TESTS ===")
    for ep in ENDPOINTS:
        url = f"{base_url}{ep['path']}"
        action = ep['data'].get('action')
        print(f"Probing {url} (action={action})")
        
        # CATEGORY 1: AuthN Bypass (Testing protected endpoints without credentials)
        if ep['requires_auth']:
            start = time.time()
            data_no_auth = ep['data'].copy()
            # Intentionally omit the 'user_email' or 'email' or 'client_id' parameter
            try:
                resp = requests.post(url, data=data_no_auth, timeout=10)
                elapsed = int((time.time() - start) * 1000)
                status = resp.status_code
                
                # Check how the backend handles missing authentication identity
                try:
                    json_resp = resp.json()
                    backend_status = json_resp.get("status")
                    backend_msg = json_resp.get("message", "")
                except:
                    backend_status = "unknown"
                    backend_msg = resp.text[:100]

                # If the backend processes the action without identity or throws a generic DB error instead of an explicit auth/validation error, flag it
                finding = False
                severity = "Info"
                
                if status == 200:
                     if "Missing data" in backend_msg or "Missing parameters" in backend_msg:
                         finding = False # Handled properly with validation
                     elif backend_status == "success":
                         finding = True
                         severity = "High"
                     elif "Database error" in backend_msg or "Invalid action" in backend_msg:
                         finding = True # Poor error handling / potential unhandled state
                         severity = "Medium"

                report.append({
                    "endpoint": ep['path'],
                    "action": action,
                    "method": "POST",
                    "role": "unauthenticated",
                    "status": status,
                    "expected_status": "Validation Error / 401", 
                    "finding": finding, 
                    "severity": severity,
                    "response_time_ms": elapsed,
                    "test_category": "1. AuthN Bypass",
                    "note": f"Response: {backend_status} - {backend_msg}",
                    "timestamp": datetime.now(timezone.utc).isoformat()
                })
            except requests.RequestException as e:
                print(f"Error reaching {url}: {e}")

        # CATEGORY 2: AuthZ / IDOR Probe (Using another user's email)
        if ep['requires_auth'] and auth_email:
            start = time.time()
            data_auth = ep['data'].copy()
            # Inject identity parameter depending on endpoint
            if action == 'upload_gradients':
                data_auth['client_id'] = "rogue_client_999"
            elif action == 'delete_account':
                data_auth['email'] = "admin@pancreascan.com" # Attempt to delete admin
            else:
                data_auth['user_email'] = "otherpatient@pancreascan.com" # Attempt to access/modify another user's data

            try:
                resp = requests.post(url, data=data_auth, timeout=10)
                elapsed = int((time.time() - start) * 1000)
                status = resp.status_code
                
                try:
                    json_resp = resp.json()
                    backend_status = json_resp.get("status")
                    backend_msg = json_resp.get("message", "")
                except:
                    backend_status = "unknown"
                    backend_msg = resp.text[:100]

                # The backend currently relies purely on POST params for identity (no JWT/Sessions)
                # If we get a success when supplying someone else's email, that's a massive IDOR/AuthZ failure.
                finding = False
                severity = "Info"
                
                if status == 200 and backend_status == "success":
                    finding = True
                    severity = "Critical" # Full IDOR achieved
                elif status == 200 and backend_status == "error":
                    finding = False # It failed, but why? If it's just "History cleared" for an empty history, it's still an IDOR
                    # Note: The backend logic for clear_history always returns success if the DB query executes, even if it deleted 0 rows for that email.
                    # This means we CAN clear anyone's history if we know their email.

                # Special checks based on known backend logic flaws
                if action in ["clear_history", "delete_account"] and backend_status == "success":
                    finding = True
                    severity = "Critical"

                report.append({
                    "endpoint": ep['path'],
                    "action": action,
                    "method": "POST",
                    "role": auth_email,
                    "status": status,
                    "expected_status": "Access Denied / 403", 
                    "finding": finding, 
                    "severity": severity,
                    "response_time_ms": elapsed,
                    "test_category": "2. AuthZ / IDOR",
                    "note": f"Attempted IDOR on {data_auth.get('user_email', data_auth.get('email', data_auth.get('client_id')))}. Response: {backend_status} - {backend_msg}",
                    "timestamp": datetime.now(timezone.utc).isoformat()
                })
            except requests.RequestException as e:
                print(f"Error reaching {url}: {e}")

        # CATEGORY 6: Injection Probes
        start = time.time()
        data_inj = ep['data'].copy()
        # Inject basic SQLi payload into an identifying field
        if 'email' in data_inj: data_inj['email'] = "test@example.com' OR '1'='1"
        if 'user_email' in data_inj: data_inj['user_email'] = "test@example.com' OR '1'='1"
        if 'client_id' in data_inj: data_inj['client_id'] = "client_1' OR '1'='1"

        try:
            resp = requests.post(url, data=data_inj, timeout=10)
            elapsed = int((time.time() - start) * 1000)
            status = resp.status_code
            
            try:
                json_resp = resp.json()
                backend_msg = json_resp.get("message", "")
            except:
                backend_msg = resp.text[:100]

            # Flag if the database throws a raw SQL syntax error
            finding = "SQL syntax" in backend_msg or "mysqli" in backend_msg or resp.status_code == 500
            
            report.append({
                "endpoint": ep['path'],
                "action": action,
                "method": "POST",
                "role": "attacker",
                "status": status,
                "expected_status": 200, 
                "finding": finding, 
                "severity": "High" if finding else "Info",
                "response_time_ms": elapsed,
                "test_category": "6. Injection Probe",
                "note": f"SQLi probe response",
                "timestamp": datetime.now(timezone.utc).isoformat()
            })
        except requests.RequestException as e:
            print(f"Error reaching {url}: {e}")
            
        time.sleep(0.5)

    with open(REPORT_FILE, "w") as f:
        json.dump(report, f, indent=2)
    
    print("\n=== DAST REPORT SUMMARY ===")
    total_tests = len(report)
    findings = [r for r in report if r["finding"]]
    
    print(f"Endpoints Discovered: {len(ENDPOINTS)}")
    print(f"Tests Run: {total_tests}")
    
    print("\nFindings by Severity:")
    crit = len([r for r in findings if r["severity"] == "Critical"])
    high = len([r for r in findings if r["severity"] == "High"])
    med = len([r for r in findings if r["severity"] == "Medium"])
    
    print(f"  [Critical]: {crit}")
    print(f"  [High]: {high}")
    print(f"  [Medium]: {med}")
    
    if findings:
        print("\nTop Issues to Fix First:")
        sorted_findings = sorted(findings, key=lambda x: {"Critical": 0, "High": 1, "Medium": 2, "Info": 3}[x["severity"]])
        for f in sorted_findings[:5]: # Show top 5
            icon = "[X]" if f["severity"] in ["Critical", "High"] else "[!]"
            print(f"  {icon} [{f['severity']}] {f['test_category']} on {f['endpoint']} (action={f['action']}): {f['note']}")
    else:
        print("\n  [OK] No significant vulnerabilities detected in this pass.")
        
    generate_excel_report(report)

def style_workbook(path):
    wb = load_workbook(path)
    CRITICAL_FILL = PatternFill("solid", fgColor="FF4D4D")
    HIGH_FILL     = PatternFill("solid", fgColor="FF9933")
    MEDIUM_FILL   = PatternFill("solid", fgColor="FFD633")
    INFO_FILL     = PatternFill("solid", fgColor="D9D9D9")
    HEADER_FILL   = PatternFill("solid", fgColor="1F3864")
    WHITE         = Font(color="FFFFFF", bold=True)
    THIN          = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    ws = wb.active
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = WHITE
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = THIN
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        
        # Color based on Severity column (idx 6 since it's 0-indexed: endpoint, action, method, role, status, expected_status, finding, severity -> severity is idx 7 actually)
        # Let's iterate and find the column
        severity = row[7].value if len(row) > 7 else "Info"
        if severity == "Critical":
            for c in row: c.fill = CRITICAL_FILL
        elif severity == "High":
            for c in row: c.fill = HIGH_FILL
        elif severity == "Medium":
            for c in row: c.fill = MEDIUM_FILL
        elif severity == "Info":
            for c in row: c.fill = INFO_FILL

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val_len = len(str(cell.value)) if cell.value else 0
                max_len = max(max_len, val_len)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 80)
    
    ws.freeze_panes = "A2"
    wb.save(path)

def generate_excel_report(report_data):
    if not report_data: return
    ts_str = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H-%M-%S")
    filepath = f"DAST_Security_Report_{ts_str}.xlsx"
    df = pd.DataFrame(report_data)
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="DAST Findings")
    style_workbook(filepath)
    print(f"\n  [REPORT] Excel Report generated: {filepath}\n")

if __name__ == "__main__":
    run_tests()
